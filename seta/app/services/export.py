"""Выгрузка в Excel и PDF: поручения, решения, встречи за период.

Выгрузка — это тот же список, что человек видит в интерфейсе, только файлом.
Поэтому строки отбираются теми же условиями видимости, что и везде: своей
логики прав здесь нет ни одной строки. Выгрузка, показывающая больше экрана, —
самый тихий способ обойти права, какой бывает.

Файл уходит в чат и на диск не кладётся: лишняя копия корпоративных данных
на сервере никому не нужна. Каждая выгрузка пишется в журнал аудита.
"""
import io
from dataclasses import dataclass
from datetime import datetime

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.timeutil import to_local
from app.models import Decision, Meeting, Task, User
from app.services import decisions as decision_service
from app.services import meetings as meeting_service
from app.services import tasks as task_service
from app.services.audit import write_audit
from app.services.decisions import STATUS_LABELS as DECISION_LABELS
from app.services.rbac import Grant, has_permission, visible_department_ids
from app.services.tasks import PRIORITY_LABELS, STATUS_LABELS as TASK_LABELS

# Шрифт с кириллицей. Базовые шрифты PDF русских букв не содержат — без него
# отчёт вышел бы из знаков вопроса, и это выяснилось бы у владельца, а не здесь.
FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
MAX_ROWS = 5000


@dataclass
class Sheet:
    """Готовая таблица: заголовок, колонки, строки."""

    title: str
    columns: list[str]
    rows: list[list[str]]

    @property
    def empty(self) -> bool:
        return not self.rows


def _when(value: datetime | None, tz: str) -> str:
    return to_local(value, tz).strftime("%d.%m.%Y %H:%M") if value else "—"


async def collect(
    session: AsyncSession,
    *,
    user: User,
    grants: dict[str, Grant],
    kind: str,
    since: datetime,
    until: datetime,
) -> Sheet:
    """Собирает данные выгрузки. Права — те же условия, что в поиске и списках."""
    visible = await visible_department_ids(session, user)
    tz = user.timezone

    if kind == "tasks":
        rows = (
            await session.execute(
                select(Task)
                .where(
                    *task_service.visible_filter(user, grants, visible),
                    Task.created_at >= since,
                    Task.created_at < until,
                )
                .order_by(Task.created_at)
                .limit(MAX_ROWS)
            )
        ).scalars().all()
        people = await _names(session, {t.assignee_id for t in rows} | {t.creator_id for t in rows})
        return Sheet(
            title="Поручения",
            columns=["№", "Поручение", "Исполнитель", "Автор", "Срок", "Приоритет", "Состояние"],
            rows=[
                [
                    str(t.id), t.title,
                    people.get(t.assignee_id, "—"), people.get(t.creator_id, "—"),
                    _when(t.due_at, tz),
                    PRIORITY_LABELS.get(t.priority, t.priority),
                    TASK_LABELS.get(t.status, t.status),
                ]
                for t in rows
            ],
        )

    if kind == "decisions":
        rows = (
            await session.execute(
                select(Decision)
                .where(
                    *decision_service.visible_filter(user, grants, visible),
                    Decision.created_at >= since,
                    Decision.created_at < until,
                )
                .order_by(Decision.created_at)
                .limit(MAX_ROWS)
            )
        ).scalars().all()
        people = await _names(
            session, {d.author_id for d in rows} | {d.responsible_id for d in rows if d.responsible_id}
        )
        return Sheet(
            title="Решения",
            columns=["№", "Решение", "Автор", "Ответственный", "Срок", "Принято", "Состояние"],
            rows=[
                [
                    str(d.id), d.title,
                    people.get(d.author_id, "—"),
                    people.get(d.responsible_id, "—") if d.responsible_id else "—",
                    _when(d.due_date, tz), _when(d.created_at, tz),
                    DECISION_LABELS.get(d.status, d.status),
                ]
                for d in rows
            ],
        )

    if kind == "meetings":
        rows = (
            await session.execute(
                select(Meeting)
                .where(
                    *meeting_service.visible_filter(user, grants, visible),
                    Meeting.start_at >= since,
                    Meeting.start_at < until,
                )
                .order_by(Meeting.start_at)
                .limit(MAX_ROWS)
            )
        ).scalars().all()
        people = await _names(session, {m.owner_id for m in rows})
        return Sheet(
            title="Встречи",
            columns=["№", "Тема", "Ведёт", "Начало", "Минут", "Состояние", "Переносов"],
            rows=[
                [
                    str(m.id), m.title, people.get(m.owner_id, "—"),
                    _when(m.start_at, tz),
                    str(int((m.end_at - m.start_at).total_seconds() // 60)),
                    m.status, str(m.reschedule_count),
                ]
                for m in rows
            ],
        )

    return Sheet(title="Неизвестно", columns=[], rows=[])


async def _names(session: AsyncSession, ids: set[int]) -> dict[int, str]:
    """Имена одним запросом: в цикле по строкам выгрузки их брать нельзя."""
    ids = {i for i in ids if i}
    if not ids:
        return {}
    rows = await session.execute(select(User.id, User.full_name).where(User.id.in_(ids)))
    return {row[0]: row[1] for row in rows.all()}


# Знаки, с которых Excel начинает читать содержимое ячейки как формулу.
FORMULA_STARTERS = ("=", "+", "-", "@", "\t", "\r")


def _append_as_text(page, row: list[str]) -> None:
    """Пишет строку так, чтобы текст от человека остался текстом.

    `openpyxl` сам решает, что строка, начинающаяся с `=`, — это формула,
    и пишет её в ячейку формулой. Поручение с названием `=cmd|'/c calc'!A1`
    стало бы исполняемой ячейкой у того, кто откроет отчёт. Это тот же случай,
    что `esc()` для сообщений: текст от человека нигде не должен становиться
    разметкой или кодом.

    Значение не искажается: тип ячейки меняется на строковый, а `quotePrefix`
    помечает её как заведомо текстовую — на случай, если содержимое скопируют
    в другую таблицу.
    """
    page.append(row)
    index = page.max_row
    for column, value in enumerate(row, start=1):
        if isinstance(value, str) and value.startswith(FORMULA_STARTERS):
            cell = page.cell(row=index, column=column)
            cell.data_type = "s"
            cell.quotePrefix = True


def to_xlsx(sheet: Sheet) -> bytes:
    """Таблица в Excel. Ширина колонок подбирается по содержимому."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    book = Workbook()
    page = book.active
    page.title = sheet.title[:31]
    page.append(sheet.columns)
    for cell in page[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    for row in sheet.rows:
        _append_as_text(page, row)

    for index, column in enumerate(sheet.columns, start=1):
        width = max([len(column)] + [len(str(r[index - 1])) for r in sheet.rows] + [8])
        page.column_dimensions[page.cell(row=1, column=index).column_letter].width = min(
            width + 2, 60
        )
    page.freeze_panes = "A2"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def to_pdf(sheet: Sheet, *, heading: str) -> bytes:
    """Таблица в PDF — для доклада наверх, где Excel неуместен."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("dejavu", "", FONT_PATH)
    pdf.add_page()
    pdf.set_font("dejavu", size=14)
    pdf.multi_cell(0, 8, heading)
    pdf.ln(2)

    usable = pdf.w - 2 * pdf.l_margin
    widths = _column_widths(sheet, usable)

    pdf.set_font("dejavu", size=8)
    pdf.set_fill_color(235, 235, 235)
    for column, width in zip(sheet.columns, widths):
        pdf.cell(width, 7, column, border=1, fill=True)
    pdf.ln()

    for row in sheet.rows:
        if pdf.get_y() > pdf.h - 20:
            pdf.add_page()
        for value, width in zip(row, widths):
            pdf.cell(width, 6, _fit(str(value), width), border=1)
        pdf.ln()

    if sheet.empty:
        pdf.ln(4)
        pdf.multi_cell(0, 6, "За выбранный период данных нет.")

    return bytes(pdf.output())


def _column_widths(sheet: Sheet, usable: float) -> list[float]:
    """Ширины по доле содержимого, но так, чтобы сумма равнялась странице."""
    if not sheet.columns:
        return []
    weights = []
    for index, column in enumerate(sheet.columns):
        longest = max([len(column)] + [len(str(r[index])) for r in sheet.rows[:200]] + [4])
        weights.append(min(longest, 60))
    total = sum(weights)
    return [usable * w / total for w in weights]


def _fit(value: str, width_mm: float) -> str:
    """Обрезает значение под ширину колонки: перенос строки ломал бы таблицу."""
    allowed = max(3, int(width_mm / 1.8))
    return value if len(value) <= allowed else value[: allowed - 1] + "…"


async def build(
    session: AsyncSession,
    *,
    user: User,
    grants: dict[str, Grant],
    kind: str,
    since: datetime,
    until: datetime,
    fmt: str = "xlsx",
) -> tuple[bytes | None, str, str | None]:
    """Готовит файл выгрузки. Возвращает (данные, имя файла, причина отказа)."""
    if not has_permission(grants, "export.read"):
        return None, "", "Нет права на выгрузку."
    if kind not in ("tasks", "decisions", "meetings"):
        return None, "", "Неизвестный вид выгрузки."

    sheet = await collect(
        session, user=user, grants=grants, kind=kind, since=since, until=until
    )
    if sheet.empty:
        # Пустой файл выглядит как сбой системы. Честный ответ полезнее.
        return None, "", "За выбранный период данных нет."

    period = (
        f"{to_local(since, user.timezone):%d.%m.%Y}—{to_local(until, user.timezone):%d.%m.%Y}"
    )
    heading = f"{sheet.title}: {period}"
    if fmt == "pdf":
        data, name = to_pdf(sheet, heading=heading), f"{kind}-{period}.pdf"
    else:
        data, name = to_xlsx(sheet), f"{kind}-{period}.xlsx"

    await write_audit(
        session, actor_id=user.id, action="export.build",
        entity_type="export", entity_id=None,
        after={"kind": kind, "format": fmt, "rows": len(sheet.rows)},
    )
    return data, name, None
